"""
Attendance Utilities
Handles CSV file operations, Excel export, and data processing
"""

import pandas as pd
import os
from datetime import datetime
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

ATTENDANCE_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'attendance_records')

def ensure_attendance_directory():
    """Ensure attendance directory exists"""
    if not os.path.exists(ATTENDANCE_DIR):
        os.makedirs(ATTENDANCE_DIR)

def get_csv_filename(subject, date=None):
    """
    Generate CSV filename for attendance records
    
    Args:
        subject: Subject name
        date: Date string (YYYY-MM-DD) or None for today
    
    Returns:
        Full path to CSV file
    """
    if date is None:
        date = datetime.now().strftime('%Y-%m-%d')
    
    filename = f"{subject}_{date}.csv"
    return os.path.join(ATTENDANCE_DIR, filename)

def save_attendance_to_csv(student_id, name, subject, date=None, time_in=None):
    """
    Save attendance record to CSV file
    
    Args:
        student_id: Student ID
        name: Student name
        subject: Subject name
        date: Date string (optional)
        time_in: Time string (optional)
    
    Returns:
        Boolean indicating success
    """
    ensure_attendance_directory()
    
    if date is None:
        date = datetime.now().strftime('%Y-%m-%d')
    if time_in is None:
        time_in = datetime.now().strftime('%H:%M:%S')
    
    csv_file = get_csv_filename(subject, date)
    
    # Check if file exists and if student already marked
    file_exists = os.path.exists(csv_file)
    
    if file_exists:
        # Check for duplicate entry
        with open(csv_file, 'r', newline='') as f:
            reader = csv.DictReader(f)
            for row in reader:
                if row['Student_ID'] == student_id:
                    return False  # Already marked
    
    # Append to CSV
    with open(csv_file, 'a', newline='') as f:
        fieldnames = ['Student_ID', 'Name', 'Subject', 'Date', 'Time_In']
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        
        if not file_exists:
            writer.writeheader()
        
        writer.writerow({
            'Student_ID': student_id,
            'Name': name,
            'Subject': subject,
            'Date': date,
            'Time_In': time_in
        })
    
    return True

def read_attendance_csv(subject, date=None):
    """
    Read attendance records from CSV file
    
    Args:
        subject: Subject name
        date: Date string (optional, if None reads all dates)
    
    Returns:
        List of attendance records
    """
    ensure_attendance_directory()
    
    if date:
        # Read specific date
        csv_file = get_csv_filename(subject, date)
        if not os.path.exists(csv_file):
            return []
        
        df = pd.read_csv(csv_file)
        return df.to_dict('records')
    else:
        # Read all CSV files for this subject
        all_records = []
        csv_files = [f for f in os.listdir(ATTENDANCE_DIR) 
                    if f.startswith(subject) and f.endswith('.csv')]
        
        for csv_file in csv_files:
            file_path = os.path.join(ATTENDANCE_DIR, csv_file)
            df = pd.read_csv(file_path)
            all_records.extend(df.to_dict('records'))
        
        return all_records

def get_all_attendance_records():
    """
    Read all attendance records from all CSV files
    
    Returns:
        List of all attendance records
    """
    ensure_attendance_directory()
    
    all_records = []
    csv_files = [f for f in os.listdir(ATTENDANCE_DIR) if f.endswith('.csv')]
    
    for csv_file in csv_files:
        file_path = os.path.join(ATTENDANCE_DIR, csv_file)
        try:
            df = pd.read_csv(file_path)
            all_records.extend(df.to_dict('records'))
        except Exception as e:
            print(f"Error reading {csv_file}: {e}")
    
    return all_records

def export_to_excel(records, subject=None):
    """
    Export attendance records to Excel file
    
    Args:
        records: List of attendance records
        subject: Subject name for filename
    
    Returns:
        Path to generated Excel file
    """
    ensure_attendance_directory()
    
    if not records:
        return None
    
    # Create DataFrame
    df = pd.DataFrame(records)
    
    # Generate filename
    if subject:
        filename = f"{subject}_attendance_{datetime.now().strftime('%Y%m%d')}.xlsx"
    else:
        filename = f"all_attendance_{datetime.now().strftime('%Y%m%d')}.xlsx"
    
    filepath = os.path.join(ATTENDANCE_DIR, filename)
    
    # Create Excel file with formatting
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance Records"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Write headers
    headers = list(df.columns)
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Write data
    for row_num, record in enumerate(df.values, 2):
        for col_num, value in enumerate(record, 1):
            ws.cell(row=row_num, column=col_num, value=value)
    
    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save workbook
    wb.save(filepath)
    
    return filepath

def get_attendance_statistics(subject=None, date=None):
    """
    Calculate attendance statistics
    
    Args:
        subject: Subject name (optional)
        date: Date string (optional)
    
    Returns:
        Dictionary with statistics
    """
    if subject:
        records = read_attendance_csv(subject, date)
    else:
        records = get_all_attendance_records()
    
    total_present = len(records)
    
    # Get unique students
    unique_students = set()
    for record in records:
        unique_students.add(record.get('Student_ID', ''))
    
    # Get subjects breakdown
    subject_breakdown = {}
    for record in records:
        subj = record.get('Subject', 'Unknown')
        if subj in subject_breakdown:
            subject_breakdown[subj] += 1
        else:
            subject_breakdown[subj] = 1
    
    # Get date breakdown
    date_breakdown = {}
    for record in records:
        dt = record.get('Date', 'Unknown')
        if dt in date_breakdown:
            date_breakdown[dt] += 1
        else:
            date_breakdown[dt] = 1
    
    return {
        'total_present': total_present,
        'unique_students': len(unique_students),
        'subject_breakdown': subject_breakdown,
        'date_breakdown': date_breakdown
    }

def get_student_attendance_history(student_id):
    """
    Get attendance history for a specific student
    
    Args:
        student_id: Student ID
    
    Returns:
        List of attendance records for the student
    """
    all_records = get_all_attendance_records()
    student_records = [r for r in all_records if r.get('Student_ID') == student_id]
    
    return student_records

def filter_attendance_by_date_range(records, start_date, end_date):
    """
    Filter attendance records by date range
    
    Args:
        records: List of attendance records
        start_date: Start date string (YYYY-MM-DD)
        end_date: End date string (YYYY-MM-DD)
    
    Returns:
        Filtered list of records
    """
    filtered = []
    
    for record in records:
        record_date = record.get('Date', '')
        if start_date <= record_date <= end_date:
            filtered.append(record)
    
    return filtered

def get_available_subjects():
    """
    Get list of all subjects that have attendance records
    
    Returns:
        List of subject names
    """
    ensure_attendance_directory()
    
    subjects = set()
    csv_files = [f for f in os.listdir(ATTENDANCE_DIR) if f.endswith('.csv')]
    
    for csv_file in csv_files:
        # Extract subject from filename (format: subject_date.csv)
        subject = '_'.join(csv_file.split('_')[:-1])
        if subject:
            subjects.add(subject)
    
    return sorted(list(subjects))

def get_available_dates(subject=None):
    """
    Get list of all dates that have attendance records
    
    Args:
        subject: Subject name (optional filter)
    
    Returns:
        List of date strings
    """
    ensure_attendance_directory()
    
    dates = set()
    csv_files = [f for f in os.listdir(ATTENDANCE_DIR) if f.endswith('.csv')]
    
    for csv_file in csv_files:
        if subject:
            if not csv_file.startswith(subject):
                continue
        
        # Extract date from filename (format: subject_YYYY-MM-DD.csv)
        parts = csv_file.replace('.csv', '').split('_')
        if len(parts) >= 2:
            date_str = parts[-1]
            dates.add(date_str)
    
    return sorted(list(dates), reverse=True)
